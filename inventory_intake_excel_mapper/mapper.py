from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd

try:
    from rapidfuzz import process as fuzz_process
except Exception:  # pragma: no cover - optional dependency
    fuzz_process = None

try:
    import numpy as np
except Exception:  # pragma: no cover - numpy comes with pandas but guard anyway
    np = None

from openpyxl import load_workbook


@dataclass
class MapperConfig:
    min_score: int = 80
    fill_value: str = "N/A"
    drop_zero_mode: str = "all"  # all | any | off
    header_scan_rows: int = 20
    append: bool = False
    key: Optional[str] = None
    source_sheet: Optional[str] = None
    target_sheet: Optional[str] = None


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, pd.Series):
        return value.map(is_blank).all()
    if isinstance(value, (list, tuple)):
        return all(is_blank(v) for v in value)
    if np is not None and isinstance(value, np.ndarray):
        return pd.isna(value).all()
    if isinstance(value, str) and not value.strip():
        return True
    if pd.isna(value):
        return True
    return False


def _best_match(target: str, candidates: Sequence[str]) -> Tuple[Optional[str], int]:
    if not candidates:
        return None, 0
    if fuzz_process is not None:
        match, score, _ = fuzz_process.extractOne(
            target, candidates, processor=normalize_header
        )
        return match, int(score)

    best = None
    best_score = 0
    target_norm = normalize_header(target)
    for candidate in candidates:
        cand_norm = normalize_header(candidate)
        if not cand_norm:
            continue
        score = _sequence_score(target_norm, cand_norm)
        if score > best_score:
            best_score = score
            best = candidate
    return best, best_score


def _sequence_score(a: str, b: str) -> int:
    from difflib import SequenceMatcher

    return int(SequenceMatcher(None, a, b).ratio() * 100)


def build_header_mapping(
    source_headers: Sequence[str],
    target_headers: Sequence[str],
    min_score: int,
) -> Dict[str, Optional[str]]:
    mapping: Dict[str, Optional[str]] = {}
    source_headers = [h for h in source_headers if normalize_header(h)]
    for target in target_headers:
        if not normalize_header(target):
            mapping[target] = None
            continue
        best, score = _best_match(target, source_headers)
        mapping[target] = best if best and score >= min_score else None
    return mapping


def load_sources(
    paths: Iterable[Path],
    sheet_name: Optional[str],
    header_scan_rows: int = 20,
) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []
    for path in paths:
        wb = load_workbook(path, data_only=True)
        names = [sheet_name] if sheet_name else wb.sheetnames
        for name in names:
            if name not in wb.sheetnames:
                raise ValueError(f"Sheet not found in {path}: {name}")
            ws = wb[name]
            header_row = detect_header_row(ws, header_scan_rows)
            header_values = [cell.value for cell in ws[header_row]]
            last_idx = None
            for idx, value in enumerate(header_values):
                if normalize_header(value):
                    last_idx = idx
            if last_idx is None:
                continue
            header_values = header_values[: last_idx + 1]
            df = pd.read_excel(
                path,
                sheet_name=name,
                header=header_row - 1,
                dtype=object,
            )
            header_values += [""] * (len(df.columns) - len(header_values))
            keep_indices = [
                idx for idx, value in enumerate(header_values) if normalize_header(value)
            ]
            df = df.iloc[:, keep_indices]
            df.columns = [str(header_values[idx]).strip() for idx in keep_indices]
            frames.append(df)

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True, sort=False)
    combined = combined.dropna(how="all")
    combined.columns = [str(c).strip() if c is not None else "" for c in combined.columns]
    combined = combined.loc[:, [c for c in combined.columns if c]]
    return combined


def detect_header_row(ws, scan_rows: int) -> int:
    best_row = 1
    best_count = 0
    for row_idx in range(1, min(ws.max_row, scan_rows) + 1):
        values = [cell.value for cell in ws[row_idx]]
        count = sum(1 for v in values if v not in (None, ""))
        if count > best_count:
            best_count = count
            best_row = row_idx
    return best_row


def sheet_to_df(ws, header_row: int, headers: List[str]) -> pd.DataFrame:
    if not headers:
        return pd.DataFrame()
    data = []
    max_col = len(headers)
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=ws.max_row,
        min_col=1,
        max_col=max_col,
        values_only=True,
    ):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        data.append(list(row))
    return pd.DataFrame(data, columns=headers)


def _is_zero(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        text = value.strip()
        return text in {"0", "0.0", "0.00"}
    if np is not None:
        if isinstance(value, (np.integer, np.floating)):
            return float(value) == 0.0
    if isinstance(value, (int, float)):
        return float(value) == 0.0
    return False


def drop_zero_rows(df: pd.DataFrame, mode: str) -> pd.DataFrame:
    if mode == "off":
        return df

    def row_is_zero(row: pd.Series) -> bool:
        values = row.tolist()
        if mode == "any":
            return any(_is_zero(v) for v in values)
        non_empty = [v for v in values if not is_blank(v)]
        if not non_empty:
            return True
        return all(_is_zero(v) for v in non_empty)

    mask = df.apply(row_is_zero, axis=1)
    return df.loc[~mask].reset_index(drop=True)


def fill_blanks(df: pd.DataFrame, fill_value: str) -> pd.DataFrame:
    df = df.copy()
    df = df.where(~df.isna(), fill_value)
    df = df.replace("", fill_value)
    return df


def merge_on_key(
    existing: pd.DataFrame, source: pd.DataFrame, key: str
) -> pd.DataFrame:
    if key not in existing.columns or key not in source.columns:
        return source
    existing = existing.set_index(key)
    source = source.set_index(key)
    merged = source.combine_first(existing)
    return merged.reset_index()


def _apply_template_style(ws, template_row: int, target_row: int, max_col: int) -> None:
    if template_row is None:
        return
    for col in range(1, max_col + 1):
        src_cell = ws.cell(row=template_row, column=col)
        tgt_cell = ws.cell(row=target_row, column=col)
        if src_cell.has_style:
            tgt_cell._style = copy(src_cell._style)
        if src_cell.number_format:
            tgt_cell.number_format = src_cell.number_format


def _to_excel_value(value: object) -> object:
    if value is pd.NA:
        return None
    if np is not None and isinstance(value, np.generic):
        return value.item()
    return value


def _build_leftover_series(source_df: pd.DataFrame, columns: Sequence[str]) -> pd.Series:
    if not columns:
        return pd.Series([pd.NA] * len(source_df), index=source_df.index)

    def first_non_blank(value: object) -> object:
        if isinstance(value, pd.Series):
            for item in value.tolist():
                if not is_blank(item):
                    return item
            return pd.NA
        if isinstance(value, (list, tuple)):
            for item in value:
                if not is_blank(item):
                    return item
            return pd.NA
        if np is not None and isinstance(value, np.ndarray):
            for item in value.tolist():
                if not is_blank(item):
                    return item
            return pd.NA
        return value

    def build_row(row: pd.Series) -> object:
        parts = []
        for col in columns:
            value = first_non_blank(row.get(col))
            if is_blank(value):
                continue
            parts.append(f"{col}: {value}")
        return " ; ".join(parts) if parts else pd.NA

    return source_df.apply(build_row, axis=1)


def apply_mapping_overrides(
    mapping: Dict[str, Optional[str]],
    source_headers: Sequence[str],
    target_headers: Sequence[str],
    target_path: Path,
) -> Dict[str, Optional[str]]:
    source_norm = {normalize_header(h): h for h in source_headers}
    target_hint = target_path.stem.lower()

    def set_if_exists(target: str, candidates: Sequence[str]) -> None:
        for candidate in candidates:
            if candidate in source_norm:
                mapping[target] = source_norm[candidate]
                return

    for target in target_headers:
        target_norm = normalize_header(target)
        if target_norm == "mpn":
            set_if_exists(
                target,
                [
                    "manufacturer pn",
                    "manufacturer part number",
                    "mfr pn",
                    "mfr part number",
                    "mfg pn",
                    "mfg part number",
                    "manufacturer part no",
                ],
            )
        if target_norm == "quantity":
            if "mob" in target_hint:
                set_if_exists(target, ["n7n30 main", "usage in n7n30 main"])
            elif "pam" in target_hint:
                set_if_exists(target, ["n7n30 pam", "usage in n7n30 pam"])
            else:
                set_if_exists(
                    target,
                    [
                        "n7n30 main",
                        "n7n30 pam",
                        "usage in n7n30 main",
                        "usage in n7n30 pam",
                    ],
                )

    return mapping


def _append_text_series(primary: pd.Series, extra: pd.Series) -> pd.Series:
    def combine(value: object, addition: object) -> object:
        if is_blank(value) and is_blank(addition):
            return pd.NA
        if is_blank(value):
            return addition
        if is_blank(addition):
            return value
        return f"{value} ; {addition}"

    return pd.Series(
        [combine(a, b) for a, b in zip(primary.tolist(), extra.tolist())],
        index=primary.index,
    )


def map_sources_to_target(
    source_df: pd.DataFrame,
    target_path: Path,
    output_path: Path,
    config: MapperConfig,
) -> None:
    if source_df.empty:
        raise ValueError("No source data loaded.")

    wb = load_workbook(target_path)
    sheet_names = [config.target_sheet] if config.target_sheet else wb.sheetnames

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        header_row = detect_header_row(ws, config.header_scan_rows)
        headers = [cell.value for cell in ws[header_row]]
        headers = [str(h).strip() if h is not None else "" for h in headers]
        while headers and not normalize_header(headers[-1]):
            headers.pop()
        headers = [
            header if header else f"Unnamed_{idx + 1}" for idx, header in enumerate(headers)
        ]
        if not headers:
            raise ValueError(f"No headers detected in sheet: {sheet_name}")

        mapping = build_header_mapping(source_df.columns.tolist(), headers, config.min_score)
        mapping = apply_mapping_overrides(
            mapping,
            source_df.columns.tolist(),
            headers,
            target_path,
        )
        note_columns = [
            header
            for header in headers
            if normalize_header(header) in {"description", "comment", "comments"}
        ]
        mapped = pd.DataFrame(index=source_df.index)
        for target in headers:
            source_col = mapping.get(target)
            if source_col in source_df.columns:
                mapped[target] = source_df[source_col]
            else:
                mapped[target] = pd.NA
        camtech_columns = [
            col for col in source_df.columns if normalize_header(col) == "camtech pn"
        ]
        if camtech_columns and note_columns:
            extra_notes = _build_leftover_series(source_df, camtech_columns)
            for note_column in note_columns:
                mapped[note_column] = _append_text_series(mapped[note_column], extra_notes)

        existing = sheet_to_df(ws, header_row, headers)
        if config.key:
            mapped = merge_on_key(existing, mapped, config.key)
        elif config.append and not existing.empty:
            mapped = pd.concat([existing, mapped], ignore_index=True)

        mapped = drop_zero_rows(mapped, config.drop_zero_mode)
        mapped = fill_blanks(mapped, config.fill_value)

        start_row = header_row + 1
        max_col = len(headers)
        template_row = start_row if ws.max_row >= start_row else None

        for row in ws.iter_rows(
            min_row=start_row,
            max_row=ws.max_row,
            min_col=1,
            max_col=max_col,
        ):
            for cell in row:
                cell.value = None

        for offset, row in enumerate(mapped.itertuples(index=False, name=None)):
            row_idx = start_row + offset
            _apply_template_style(ws, template_row, row_idx, max_col)
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=_to_excel_value(value))

    wb.save(output_path)
